import pandas as pd
import requests
import json
import os
from dotenv import load_dotenv
from openpyxl import load_workbook


load_dotenv()

POCKETBASE_URL = os.getenv("POCKETBASE_URL")
COLLECTION_NAME = "CustomerGPS_Update"
GOOGLE_MAPS_API_KEY = os.getenv("GOOGLE_MAPS_API_KEY")
FILE_PATH = "Customers database UPDATE.xlsx"

# new: detect highlighted (yellow) rows in the Excel file and return pandas indices to ignore
def get_highlighted_row_indices(file_path, sheet_name=None, header_row=1):
	"""
	Return a set of pandas DataFrame indices that correspond to rows highlighted in yellow.
	Assumes header_row is the Excel header row number (1-based). For default pandas.read_excel,
	header_row=1 (first row is header), so pandas index for Excel row N is N - header_row - 1.
	"""
	wb = load_workbook(file_path, data_only=True)
	ws = wb[sheet_name] if sheet_name else wb.active

	ignored_excel_rows = set()
	# common yellow ARGB values in Excel (match uppercased)
	yellow_signatures = ("FFFF00", "FFFFFF00")

	for row in ws.iter_rows(min_row=header_row + 1):
		for cell in row:
			rgb = None
			try:
				fg = cell.fill.fgColor
				if fg is not None:
					rgb = getattr(fg, "rgb", None)
			except Exception:
				rgb = None

			if rgb:
				rgb_up = str(rgb).upper()
				# check common signatures; adjust if your sheet uses a different yellow
				if any(sig in rgb_up for sig in yellow_signatures):
					ignored_excel_rows.add(cell.row)
					break

	# convert excel row numbers to pandas indices
	ignored_indices = {r - header_row - 1 for r in ignored_excel_rows}
	return ignored_indices

# read Excel and filter out highlighted rows
df = pd.read_excel(FILE_PATH)
df = df.where(pd.notna(df), None)

ignored_indices = get_highlighted_row_indices(FILE_PATH, header_row=1)
if ignored_indices:
	df = df[~df.index.isin(ignored_indices)]

records = df[df['CUSTOMER_NAME'] != 'None'].to_dict(orient="records")

def get_coordinates(address):
	url = f"https://maps.googleapis.com/maps/api/geocode/json"
	params = {
		'address': address,
		'key': GOOGLE_MAPS_API_KEY
	}
	try:
		response = requests.get(url, params=params)
		if response.status_code == 200:
			data = response.json()
			if data['status'] == 'OK':
				location = data['results'][0]['geometry']['location']
				return location['lat'], location['lng']
			else:
				print(f"Geocoding Error: {data['status']}")
				return None, None
		else:
			print(f"HTTP Error: {response.status_code}")
			return None, None
	except requests.exceptions.RequestException as e:
		print(f"Request Exception: {e}")
		return None, None

def upload_to_pocketbase(records):
	headers = {"Content-Type": "application/json"}

	for record in records:
		customer_name = record.get("CUSTOMER_NAME")
		full_address = record.get("CUSTOMER_FULL_ADDRESS")
		code = record.get("CUSTOMER_CODE")

		if customer_name and customer_name != 'None':
			if full_address:
				latitude, longitude = get_coordinates(full_address)
				record['latitude'] = latitude
				record['longitude'] = longitude

			if code:
				record['id'] = str(code)

			try:
				response = requests.post(
					f"{POCKETBASE_URL}/api/collections/{COLLECTION_NAME}/records",
					json=record,
					headers=headers
				)

				if response.status_code == 200:
					print(f"Uploaded: {record}")
				else:
					print(f"Error: {response.text}")

			except json.JSONDecodeError as e:
				print(f"JSON Error: {e}")
			except requests.exceptions.RequestException as e:
				print(f"Request Error: {e}")

upload_to_pocketbase(records)
